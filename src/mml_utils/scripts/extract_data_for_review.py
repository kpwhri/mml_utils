"""
Extract CUIs and text strings for some feature (e.g., fever, diarrhea, etc.) from MML output.
These can then be used to review the results to determine how useful the CUIs/text strings are.

Runs on Metamaplite output directories (text file and jsonl/mmi expected in same directory).

Expects a feature directory:
* {feature_name}.string.txt
    * unique text string to look for on each line
    * Examples:
        * BM frequent
        * BM loose
* {feature_name}.cui.txt
    * unique CUI to look for on each line
    * additional data can be included in tab-separated format
    * Examples:
        * C0687681
        * C0687681\tFeeling feverish


# Generate Reviews for Output Data

```python
import pandas as pd
from loguru import logger

SAMPLE_SIZE = 50
df = pd.GET_METADATA()  # e.g., with studyid and docid
with pd.ExcelWriter(outpath / f'features.review{SAMPLE_SIZE}.xlsx') as writer:
    logger.info(f"All notes dataset -> Unique Notes: {df['note_id'].nunique()}; Unique Patients: {df['studyid'].nunique()}")
    for file in base_path.glob('*.review.csv'):
        feature_name = file.stem.split('.')[0]
        logger.info(f'Reading feature {feature_name}: {file}')
        feature_df = pd.read_csv(file)
        logger.info(f"Source feature dataset -> Unique Notes: {feature_df['docid'].nunique()}")
        mdf = pd.merge(df, feature_df, on='docid')
        logger.info(f"Merged dataset -> Unique Notes: {mdf['docid'].nunique()}; Unique Patients: {mdf['studyid'].nunique()}")
        # output options
        mdf['Is this term referring to the feature?'] = ''
        mdf['Does the patient have this feature?'] = ''
        mdf['Optional Comments'] = ''
        unique_notes = mdf['docid'].unique()
        sampled = random.sample(sorted(unique_notes), SAMPLE_SIZE)
        rdf = mdf[mdf.docid.isin(sampled)]
        rdf = rdf[[
            'id', 'studyid', 'docid', 'precontext', 'keyword', 'postcontext', 'Is this term referring to the feature?',
            'Does the patient have this feature?', 'Optional Comments', 'fullcontext'
        ]]
        rdf.to_csv(outpath / f'{feature_name}.sample{SAMPLE_SIZE}.csv', index=False)
        rdf.to_excel(writer, sheet_name=feature_name, index=False)
```
"""
import copy
import csv
import datetime
import pathlib
import random
import re
from collections import defaultdict

import click
from loguru import logger

from mml_utils.parse.parser import extract_mml_data


def get_feature_names_from_directory(target_path: pathlib.Path):
    done = set()
    for f in target_path.iterdir():
        if not f.name.endswith('cui.txt'):
            continue
        feature_name, label, ext = f.name.split('.')
        if feature_name in done:
            continue
        yield feature_name
        done.add(feature_name)


def finditer(target, text: str, offset=0):
    """

    :param target:
    :param text:
    :param offset:
    :return: start, end
    """
    sub_offset = 0
    while True:
        try:
            index = text.index(target, sub_offset)
        except ValueError:
            return
        end = index + len(target)
        yield index + offset, end + offset
        sub_offset = end


def find_target_text(text, target, start, end):
    if target == text[start:end]:
        return start, end
    for width in [50, 100]:
        curr_text = text[max(start - width, 0):end + width]
        if target in curr_text:
            best_distance = width  # default to max
            best_start = None
            best_end = None
            for _start, _end in finditer(target, curr_text, offset=start - width):
                new_distance = min(abs(_start - start), abs(_end - start), abs(_start - end), abs(_end - end))
                if new_distance < best_distance:
                    best_distance = new_distance
                    best_start = _start
                    best_end = _end
            return best_start, best_end
    else:
        raise ValueError(f'Unable to find target `{target}` in text at {start}:{end}.')


def clean_text(text: str):
    return text.replace('\n', '\\n').replace('\t', ' ')


@click.command()
@click.argument('note-directories', nargs=-1, type=click.Path(exists=True, path_type=pathlib.Path))
@click.option('--target-path', type=click.Path(exists=True, path_type=pathlib.Path),
              help='Directory contain files like "fever.cui.txt" and "fever.string.txt"')
@click.option('--mml-format', type=str, default='json',
              help='MML Output format to look for (e.g., "json" or "mmi"); currently only supports "json" and "mmi".')
@click.option('--text-extension', type=str, default='',
              help='Add ".txt" if text files have an extension.')
@click.option('--text-encoding', type=str, default='utf8',
              help='Format to read text files into metamaplite.')
@click.option('--text-errors', type=str, default='replace',
              help='Passed to "errors" in "open" function to open file.')
@click.option('--add-cr', type=bool, default=False, is_flag=True,
              help='Add back carriage return; likely required if MML run on Windows.')
@click.option('--sample-size', type=int, default=50,
              help='Sample size to pull for each output.')
@click.option('--metadata-file', type=click.Path(exists=True, path_type=pathlib.Path), default=None,
              help='Metadata file to add additional columns to output:'
                   ' HEADER = note_id, studyid, date, etc.; VALUES = 1, A2E, 05/06/2011, etc.')
@click.option('--replacements', type=str, multiple=True,
              help='Replace text to fix offset issues. Arguments should look like "from==to" which will'
                   ' replace "from" with "to" before checking offsets.')
def _extract_data_for_review(note_directories: list[pathlib.Path], target_path: pathlib.Path = pathlib.Path('.'),
                             mml_format='json', text_extension='', text_encoding='utf8',
                             text_errors='replace', add_cr=False,
                             sample_size=50, metadata_file=None,
                             replacements=None):
    extract_data_for_review(note_directories, target_path, mml_format, text_extension, text_encoding,
                            text_errors=text_errors, add_cr=add_cr, sample_size=sample_size,
                            metadata_file=metadata_file, replacements=replacements)


def extract_data_for_review(note_directories: list[pathlib.Path], target_path: pathlib.Path = pathlib.Path('.'),
                            mml_format='json', text_extension='', text_encoding='utf8',
                            text_errors='replace', add_cr=False, sample_size=50, metadata_file=None,
                            replacements=None):
    """

    :param text_errors:
    :param add_cr:
    :param sample_size:
    :param metadata_file:
    :param replacements:
    :param text_encoding:
    :param note_directories:
    :param target_path:
    :param mml_format:
    :param text_extension: must include `.` (e.g., `.txt`)
    :return:
    """
    if sample_size:
        try:
            import openpyxl
        except ImportError:
            logger.warning(f'Unable to import openpyxl to build review sets:'
                           f' run `pip install openpyxl` if you want Excel files rather than CSV files to review.')

    outpath = target_path / f'review_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}'
    outpath.mkdir(exist_ok=False)
    note_ids = defaultdict(list)
    for feature_name in get_feature_names_from_directory(target_path):
        target_cuis = load_first_column(target_path / f'{feature_name}.cui.txt')
        logger.info(f'Starting "{feature_name}" with {len(target_cuis)} target CUIs.')
        # create pattern: sort by longest to ensure longest regex is matched first
        target_regex = build_regex_from_file(target_path, feature_name)
        unique_id = 0
        with open(outpath / f'{feature_name}.review.csv', 'w',
                  newline='', encoding=text_encoding) as fh:
            writer = csv.DictWriter(
                fh,
                fieldnames=['id', 'note_id', 'start', 'end', 'length', 'negation', 'type',
                            'precontext', 'keyword', 'postcontext', 'fullcontext']
            )
            writer.writeheader()
            for note_directory in note_directories:
                logger.info(f'Reading directory {note_directory}.')
                note_count = 0
                no_text_file_count = 0
                for mml_file in note_directory.glob(f'*.{mml_format}'):
                    txt_file = note_directory / f'{mml_file.stem}{text_extension}'
                    if not txt_file.exists():
                        logger.warning(f'Failed to find corresponding text file:'
                                       f' {txt_file.name} (extension: {text_extension}).')
                        no_text_file_count += 1
                        continue
                    note_count += 1
                    with open(txt_file, encoding=text_encoding, errors=text_errors) as fh:
                        text = fh.read()
                    if add_cr:
                        text = text.replace('\n', '\r\n')
                    if replacements:
                        for _from, _to in replacements:
                            text = text.replace(_from, _to)
                    cui_data = []
                    for data in extract_mml_data(mml_file, target_cuis=target_cuis, output_format=mml_format):
                        try:
                            start, end = find_target_text(text, data['matchedtext'], data['start'], data['end'])
                        except ValueError as ve:
                            logger.error(f'Lookup failed for {mml_file}.')
                            logger.exception(ve)
                            raise
                        if cui_data and (cui_data[-1][0] <= start <= cui_data[-1][1]):  # overlaps?
                            # keep longer if overlaps with previous cui
                            if end - start > cui_data[-1][1] - cui_data[-1][0]:
                                cui_data[-1] = start, end, True, data['negated']
                        else:  # no overlap
                            cui_data.append((start, end, True, data['negated']))
                    # find mentions from string
                    text_data = []
                    for m in target_regex.finditer(text):
                        start, end = m.start(), m.end()
                        overlap = False
                        for cui_start, cui_end, is_cui, negated in cui_data:
                            if not is_cui or cui_start > start:
                                break
                            # is there overlap from cuis?
                            if is_cui and start <= cui_start <= end or start <= cui_end <= end:
                                overlap = True
                                break  # skip if cui already contained
                        if not overlap:
                            text_data.append((start, end, False, -1))
                    # output remaining matches
                    prev_unique_id = unique_id
                    for start, end, is_cui, negated in sorted(cui_data + text_data):
                        writer.writerow({
                            'id': unique_id,
                            'note_id': mml_file.stem,
                            'start': start,
                            'end': end,
                            'length': end - start,
                            'negation': negated,
                            'type': 'CUI' if is_cui else 'TEXT',
                            'precontext': clean_text(text[max(0, start - 100):start]),
                            'keyword': clean_text(text[start:end]),
                            'postcontext': clean_text(text[end:end + 100]),
                            'fullcontext': clean_text(text[max(0, start - 250): end + 250]),
                        })
                        unique_id += 1
                    # record all note ids by feature for sampling
                    if sample_size and prev_unique_id < unique_id:
                        note_ids[feature_name].append(mml_file.stem)
                logger.info(f'Completed processing {note_count} notes (Total Matches: {unique_id}).')
                if no_text_file_count:
                    logger.warning(f'Failed to find {no_text_file_count} text files.')
    if sample_size:
        compile_to_excel(outpath, note_ids, text_encoding, sample_size, metadata_file)


def build_regex_from_file(target_path, feature_name):
    return build_regex(
        load_first_column(target_path / f'{feature_name}.string.txt')
    )


def build_regex(term_list):
    target_pattern = re.sub(
        r'\s+', r'\\s+',
        '|'.join(re.escape(x) for x in sorted(term_list, key=lambda x: -len(x)))
    )
    target_regex = re.compile(f"({target_pattern})", re.I)
    return target_regex


def load_first_column(file: pathlib.Path):
    res = {}
    with open(file) as fh:
        for line in fh:
            if line.strip():
                res[line.strip().split('\t')[0]] = 1
    return res


@click.command()
@click.option('--outpath', type=click.Path(exists=True, path_type=pathlib.Path),
              help='Directory contain files like "fever.review.csv"')
@click.option('--text-encoding', type=str, default='utf8',
              help='Format that was used to read text files into metamaplite.')
@click.option('--sample-size', type=int, default=50,
              help='Sample size to pull for each output.')
@click.option('--metadata-file', type=click.Path(exists=True, path_type=pathlib.Path), default=None,
              help='Metadata file to add additional columns to output:'
                   ' HEADER = note_id, studyid, date, etc.; VALUES = 1, A2E, 05/06/2011, etc.')
@click.option('--build-csv', type=bool, default=False, is_flag=True,
              help='Always build CSVs (possibly in addition to Excel)')
def _compile_to_excel(outpath: pathlib.Path, text_encoding='utf8', sample_size=50, metadata_file=None, build_csv=False):
    # get note_ids to sample from
    note_ids = defaultdict(set)
    for file in outpath.glob('*.review.csv'):
        feature_name = file.stem.split('.')[0]
        with open(file, newline='', encoding=text_encoding) as fh:
            for row in csv.DictReader(fh):
                note_ids[feature_name].add(row['note_id'])
    compile_to_excel(outpath, note_ids, text_encoding=text_encoding,
                     sample_size=sample_size, metadata_file=metadata_file, build_csv=build_csv)


def compile_to_excel(outpath: pathlib.Path, note_ids, text_encoding='utf8', sample_size=50,
                     metadata_file=None, build_csv=False):
    """

    :param note_ids:
    :param outpath:
    :param text_encoding:
    :param sample_size: number of notes to sample
    :param metadata_file: csv file with:
        HEADER = note_id, studyid, date, etc.
        VALUES = 1, A2E, 05/06/2011, etc.
    :return:
    """
    # e.g., HEADER = note_id, studyid, date, etc.
    # e.g., VALUES = 1, A2E, 05/06/2011
    metadata_lkp = {}
    new_fields = []
    if metadata_file:
        with open(metadata_file, newline='') as fh:
            reader = csv.DictReader(fh)
            new_fields = set(reader.fieldnames) - {'note_id'}
            for row in reader:
                note_id = row['note_id']
                del row['note_id']
                metadata_lkp[note_id] = row

    # select random sample
    new_note_ids = {}
    for feature_name in note_ids:
        if sample_size:
            new_note_ids[feature_name] = set(
                random.sample(note_ids[feature_name], min([sample_size, len(note_ids[feature_name])]))
            )
        else:
            new_note_ids = None


    has_excel = False
    try:
        import openpyxl

        has_excel = True
    except ImportError:
        logger.warning(f'Openpyxl not installed: not building excel workbook, just CSV files.')
        logger.info(f'To build Excel workbook, install openpyxl: `pip install openpyxl` and re-run.')

    if has_excel:
        _build_excel_review_set(outpath, new_fields, new_note_ids, metadata_lkp, sample_size, text_encoding)
    if not has_excel or build_csv:
        logger.info(f'Building CSV file output.')
        _build_csv_review_set(outpath, new_fields, new_note_ids, metadata_lkp, sample_size, text_encoding)


def _build_csv_review_set(outpath, new_fields, note_ids, metadata_lkp, sample_size, text_encoding):
    for file in outpath.glob('*.review.csv'):
        feature_name = file.name.split('.')[0]
        sample_note_ids = note_ids[feature_name] if note_ids else None
        with open(file, newline='', encoding=text_encoding) as fh:
            reader = csv.DictReader(fh)
            with open(outpath / f'{feature_name}.sample{sample_size}.csv', 'w',
                      newline='', encoding=text_encoding) as out:
                writer = csv.DictWriter(out, fieldnames=reader.fieldnames[:7] + new_fields + reader.fieldnames[7:])
                writer.writeheader()
                for row in reader:
                    if sample_note_ids and row['note_id'] not in sample_note_ids:
                        continue
                    writer.writerow(row | metadata_lkp.get(row['note_id'], dict()))


def _build_excel_review_set(outpath, new_fields, note_ids, metadata_lkp, sample_size, text_encoding):
    """

    :param outpath:
    :param new_fields:
    :param note_ids:
    :param sample_size:
    :param text_encoding:
    :return:
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    # column data
    pre_columns = ['id', 'note_id', 'start', 'end', 'length', 'negation', 'type']
    post_columns = ['precontext',
                    'keyword',
                    'postcontext',
                    'Is this term referring to the feature?',
                    'Does the patient have this feature?',
                    'Optional Comments',
                    'fullcontext',
                    ]
    column_widths = {
                        'id': 12,
                        'note_id': 12,
                        'start': 12,
                        'end': 12,
                        'length': 12,
                        'negation': 12,
                        'type': 12,
                    } | {
                        field: 12 for field in new_fields
                    } | {
                        'precontext': 40,
                        'keyword': 24,
                        'postcontext': 40,
                        'Is this term referring to the feature?': 24,
                        'Does the patient have this feature?': 24,
                        'Optional Comments': 24,
                        'fullcontext': 60,
                    }
    column_alignments = [
                            [True, 'right', 'center'],
                            [True, 'right', 'center'],
                            [True, 'right', 'center'],
                            [True, 'right', 'center'],
                            [True, 'right', 'center'],
                            [True, 'right', 'center'],
                            [True, 'right', 'center'],
                        ] + [
                            [True, 'right', 'center'] for _ in new_fields
                        ] + [
                            [True, 'right', 'center'],
                            [True, 'center', 'center'],
                            [True, 'left', 'center'],
                            [True, 'center', 'center'],
                            [True, 'center', 'center'],
                            [True, 'left', 'center'],
                            [True, 'left', 'center'],
                        ]
    wb = Workbook()
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name=f'TableStyleMedium9', showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    for file in outpath.glob('*.review.csv'):
        feature_name = file.name.split('.')[0]
        sample_note_ids = note_ids[feature_name]
        max_row_index = 1  # 1-indexed
        ws = wb.create_sheet(title=f'{feature_name}')
        with open(file, newline='', encoding=text_encoding) as fh:
            reader = csv.DictReader(fh)
            max_col_letter = get_column_letter(len(column_alignments))
            ws.append(list(column_widths.keys()))
            for row in reader:
                if sample_note_ids and row['note_id'] not in sample_note_ids:
                    continue
                if row['note_id'] in metadata_lkp:
                    md = list(metadata_lkp[row['note_id']].values())
                else:
                    md = [''] * len(new_fields)
                ws.append(
                    [row[k] for k in pre_columns]
                    + md
                    + [row.get(k, '') for k in post_columns]
                )
                max_row_index += 1

            table = Table(
                displayName=feature_name,
                ref=f'A1:{max_col_letter}{max_row_index}',
                tableStyleInfo=style,
            )
            ws.add_table(table)
            for i, width in enumerate(column_widths.values(), start=1):
                ws.column_dimensions[get_column_letter(i)].width = width
            for sheet_row in ws.iter_rows():
                for i, cell in enumerate(sheet_row):
                    alignment = copy.copy(cell.alignment)
                    alignment.wrapText = column_alignments[i][0]
                    alignment.horizontal = column_alignments[i][1]
                    alignment.vertical = column_alignments[i][2]
                    cell.alignment = alignment
    wb.remove(wb[wb.sheetnames[0]])  # remove default sheet
    wb.save(outpath / f'features.review.sample{sample_size}.xlsx')


if __name__ == '__main__':
    _extract_data_for_review()
